from OneDataOfAcc import OneDataOfAcc
from OneKeyin import OneKeyin
from _231230a import data_acc_sunday, data_keyin_sunday, dict_date_subpoena, exportCompare


import linque as lq
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


import typing as t
from datetime import datetime


class ExportCompare:
    def __init__(self):
        self.row: int
        self.sh: Worksheet
        self.keyBool_acc: t.Dict[str, bool]
        self.keyBool_keyin: t.Dict[datetime, bool]
    def _init_keyBool_to_false(self,data_acc_sunday: t.List[OneDataOfAcc], data_keyin_sunday: t.List[OneKeyin]):
        # 列出所有 key, 最後再把沒輸出的輸出出來
        keyBool_acc = lq.linque(data_acc_sunday).select(lambda a1: a1.subpoena).distinct().to_dict(lambda a1: a1, lambda a1: False)
        keyBool_keyin = lq.linque(data_keyin_sunday).select(lambda a1: a1.date).distinct().to_dict(lambda a1: a1, lambda a1: False)

        self.keyBool_acc = keyBool_acc
        self.keyBool_keyin = keyBool_keyin
    def _find_pair_sunday_and_write_each_pair(self, acc: OneDataOfAcc, keyins: t.List[OneKeyin]):
        # 以 keyin 鍵為主，從 acc 中找，如果沒找到，就記載
        for a1 in lq.linque(data_keyin_sunday).group(lambda a1: a1.date):
            date: datetime = a1[0]
            r1: lq.Linque = a1[1]
            if date in dict_date_subpoena:
                subpoena = dict_date_subpoena[date]
                self.keyBool_acc[subpoena] = True
                self.keyBool_keyin[date] = True

                exportCompare.do_one_pair(subpoena, date)
            else:
                # self.keyBool_keyin[date] 會保持 False，第二階段再輸出
                pass

        pass
    def main(self, sh: Worksheet, data_acc_sunday: t.List[OneDataOfAcc], data_keyin_sunday: t.List[OneKeyin]):
        """ 主流程

        Args:
            sh (Worksheet): 指定 Sheet，產生資料在哪。
            data_acc_sunday (t.List[OneDataOfAcc]): 正航傳票，星期日的，收入的。(但沒有合併，就是相同的科目，部門，合併)
            data_keyin_sunday (t.List[OneKeyin]): 手動輸入的，非轉帳的。有合併，同科目，同部門的。
        """
        self.sh = sh
        self.row = 2

        self._init_keyBool_to_false(data_acc_sunday, data_keyin_sunday)

        self._write_header()

        self._find_pair_sunday_and_write_each_pair(data_acc_sunday, data_keyin_sunday)

        self._do_residue(data_acc_sunday, data_keyin_sunday)

        # 修正格式 (非必要)
        self._set_date_format(sh)
        self._set_money_format(sh)

    def _set_date_format(self, sh: Worksheet):
        # B, H 日期
        for i in range(2, sh.max_row+1):
            cell = sh.cell(row=i, column=2)
            cell.number_format = 'm/d'
            cell = sh.cell(row=i, column=8)
            cell.number_format = 'm/d'
    def _set_money_format(self, sh: Worksheet):
        # E, K 金額，加上千分位
        for i in range(2, sh.max_row+1):
            cell = sh.cell(row=i, column=5)
            cell.number_format = '#,##0'
            cell = sh.cell(row=i, column=11)
            cell.number_format = '#,##0'
    def _write_header(self):
        sh = self.sh

        sh.cell(row=1, column=1).value = '傳票號碼'
        sh.cell(row=1, column=2).value = '日期'
        sh.cell(row=1, column=3).value = '科目'
        sh.cell(row=1, column=4).value = '部門'
        sh.cell(row=1, column=5).value = '金額'
        sh.cell(row=1, column=6).value = '摘要'
        sh.cell(row=1, column=7).value = ''
        sh.cell(row=1, column=8).value = '日期'
        sh.cell(row=1, column=9).value = '科目'
        sh.cell(row=1, column=10).value = '部門'
        sh.cell(row=1, column=11).value = '金額'
        sh.cell(row=1, column=12).value = '摘要'

    def _chk_is_fit_any_keyin(self, acc: OneDataOfAcc, keyins: t.List[OneKeyin])->bool:
        '''其中一個 acc, 是否在 keyins 中有找到, 為了上色(綠色，表示有)
        - 科目，部門，金額要對。
        '''
        for a1 in keyins:
            if a1.money != acc.money:
                continue
            r1 = a1.subjectNumber
            if r1[0]!=acc.subject or r1[1]!=acc.department:
                continue
            return True
        return False
    def _chk_is_fit_any_acc(self, keyin: OneKeyin, accs: t.List[OneDataOfAcc])->bool:
        '''其中一個 keyin, 是否在 accs 中有找到, 為了上色(綠色，表示有)
        - 科目，部門，金額要對。
        '''
        for a1 in accs:
            if a1.money != keyin.money:
                continue
            if a1.subject!=keyin.subjectNumber[0] or a1.department!=keyin.subjectNumber[1]:
                continue
            return True
        return False

    def do_one_pair(self,acckey: str, keyinkey: datetime):
        r1: t.List[OneDataOfAcc] = lq.linque(data_acc_sunday).where(lambda a1: a1.subpoena==acckey).to_list()
        r2: t.List[OneKeyin] = lq.linque(data_keyin_sunday).where(lambda a1: a1.date==keyinkey).to_list()

        # acc 傳票號碼 日期 科目 部門 金額 摘要 ; ; keyin 日期 科目 部門 金額 摘要
        countRow = max(len(r1), len(r2))
        row = self.row
        sh = self.sh
        for i in range(countRow):
            acc = r1[i] if i<len(r1) else None
            keyin = r2[i] if i<len(r2) else None
            if acc is not None:
                self._print_acc(acc, row)
                if self._chk_is_fit_any_keyin(acc, r2):
                    cell = sh.cell(row=row, column=5)
                    # 文字顏色 綠色
                    cell.font = openpyxl.styles.Font(color='208800')

            if keyin is not None:
                self._print_keyin(keyin, row)
                if self._chk_is_fit_any_acc(keyin, r1):
                    cell = sh.cell(row=row, column=11)
                    # 文字顏色 綠色
                    cell.font = openpyxl.styles.Font(color='008800')

            row += 1
            self.row += 1
        self.row += 1 # 空白一行
        pass
    def _print_acc(self, acc: OneDataOfAcc, row: int):
        sh = self.sh

        sh.cell(row=row, column=1).value = acc.subpoena
        sh.cell(row=row, column=2).value = acc.date
        sh.cell(row=row, column=3).value = acc.subject
        sh.cell(row=row, column=4).value = acc.department
        sh.cell(row=row, column=5).value = acc.money
        sh.cell(row=row, column=6).value = acc.memo

    def _print_keyin(self, keyin: OneKeyin, row: int):
        sh = self.sh

        sh.cell(row=row, column=8).value = keyin.date
        sh.cell(row=row, column=9).value = keyin.subjectNumber[0]
        sh.cell(row=row, column=10).value = keyin.subjectNumber[1]
        sh.cell(row=row, column=11).value = keyin.money
        sh.cell(row=row, column=12).value = keyin.memo


    def _do_residue(self, data_acc_sunday: t.List[OneDataOfAcc], data_keyin_sunday: t.List[OneKeyin]):
        """ 將所有剩的，組點一整份 List """
        row = self.row

        # 剩下沒有輸出的 acc
        acc2: t.List[OneDataOfAcc] = []
        for a1 in lq.linque(self.keyBool_acc).where(lambda a1: self.keyBool_acc[a1]==False):
            r1 = lq.linque(data_acc_sunday).where(lambda a2: a2.subpoena==a1).to_list()
            acc2.extend(r1)

        # 剩下沒有輸出的 keyin
        keyin2: t.List[OneKeyin] = []
        for a1 in lq.linque(self.keyBool_keyin).where(lambda a1: self.keyBool_keyin[a1]==False):
            r1 = lq.linque(data_keyin_sunday).where(lambda a2: a2.date==a1).to_list()
            keyin2.extend(r1)

        # 輸出    
        countRow = max(len(acc2), len(keyin2))

        for i in range(countRow):
            acc = acc2[i] if i<len(acc2) else None
            keyin = keyin2[i] if i<len(keyin2) else None

            if acc is not None:
                self._print_acc(acc, row)

            if keyin is not None:
                self._print_keyin(keyin, row)

            row += 1
        self.row = row
        pass