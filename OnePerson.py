from __future__ import annotations
#%%
from openpyxl.worksheet.worksheet import Worksheet


import typing as t


class OnePerson:
    def __init__(self,id2: str,name: str, no: t.Optional[int]) -> None:
        """ 上傳奉獻資料時，只需要會友編號和姓名，其他欄位都不需要。
        ### Args:
        - id2 (str): 202301091205536 那一串
        - name (str): 陳小明/不具名
        - no (int?): 袋號
        ### 設計
        - 為何需要 no。這是因為輸入資料，可能是用奉獻袋號，所以型成 dict 過程會用到。
        - 為何 id2 不稱作id，因為欄位中還有一個資料庫的主鍵 id。
        """
        self.id2 = id2
        self.name = name
        self.no = no
    def __repr__(self) -> str:
        if self.no is None:
            return f'{self.name}'
        return f'{self.no} {self.name}'
    @property
    def noForOrder(self)->int:
        """ 提供給排序用，若有袋號，則以袋號。若無袋號，愈短的字，愈前面。接著，再以姓名第一個字的 unicode 碼。
        """
        if self.no is None:
            return 9999 + 10000 * len(self.name) + ord(self.name[0])
        return self.no
    @staticmethod
    def generate_from_excel(sh: Worksheet,row: int)->OnePerson:
        # name:2, idd2:5, no:18
        name = sh.cell(row=row, column=2).value
        id2 = sh.cell(row=row, column=5).value
        no = sh.cell(row=row, column=18).value
        if no is not None and isinstance(no, str):
            assert str.isdigit(no)
            no = int(no)
        return OnePerson(id2, name, no)