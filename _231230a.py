#%%
# 取得所有資料，兩個都會用到
from ExportCompare import ExportCompare
from get_all_data_of_acc import get_all_data_of_acc
from get_all_data_of_keyin import get_all_data_of_keyin
from write_transfer_sheet import write_transfer_sheet

data_acc = get_all_data_of_acc('2023傳票_20240109匯出.xlsx')
data_keyin_sunday, data_keyin_transfer = get_all_data_of_keyin('2023 奉獻輸入資料.xlsx')

#%%
# 產生 主日 的比較
from filter_subpoena_sunday_and_income import filter_subpoena_not_sunday_and_income, filter_subpoena_sunday_and_income
from sort_acc import sort_acc
from sort_keyin import sort_keyin
from grouping_keyin import grouping_keyin

# 產生 主日 的比較
data_acc_sunday = sort_acc( filter_subpoena_sunday_and_income(data_acc) )
data_keyin_sunday = sort_keyin( grouping_keyin(data_keyin_sunday) )

# 產生 非主日/轉帳 的比較
data_not_sunday_acc = sort_acc(filter_subpoena_not_sunday_and_income(data_acc))
deta_not_sunday_keyin = sort_keyin( data_keyin_transfer )

#%
import openpyxl
wb = openpyxl.Workbook()
wb.remove(wb["Sheet"])

sh = wb.create_sheet('主日_比較')
exportCompare = ExportCompare()
exportCompare.main(sh, data_acc_sunday, data_keyin_sunday)

sh = wb.create_sheet('轉帳_比較')
write_transfer_sheet(sh, data_not_sunday_acc, deta_not_sunday_keyin)

#%
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
import linque as lq
from OneKeyin import OneKeyin
import typing as t

sh = wb.create_sheet('正航_比較')
def fn_group1(a1: OneKeyin):    
    r1 = a1.subjectNumber
    return r1[0] # 科目
def fn_sort1(a1: t.Tuple[str,lq.Linque]):
    if str.isdigit(a1[0]):
        return int(a1[0])
    return 5000000 + ord(a1[0][0])
row = 2

def write_header1(row):
    ce = sh.cell(row=row, column=1)
    ce.value = '科目'
    ce = sh.cell(row=row, column=2)
    ce.value = '金額'
    ce = sh.cell(row=row, column=3)
    ce.value = '注：此部分請手動比對匯出的損益表_收入'
write_header1(1)

for a1 in lq.linque(data_keyin_sunday).concat(data_keyin_transfer).group(fn_group1).sort(fn_sort1):
    r1: lq.Linque = a1[1]
    ce = sh.cell(row=row, column=1)
    ce.value = a1[0]
    ce: Cell = sh.cell(row=row, column=2)    
    ce.value = r1.sum(lambda a2: a2.money)
    ce.number_format = '#,##0'
    
    row += 1

row+=1

def fn_group2(a1: OneKeyin):
    r1 = a1.subjectNumber
    return (r1[1], r1[0]) # 部門，科目
def fn_sort2(a1: t.Tuple[t.Tuple[str,str],lq.Linque]):
    key1 = a1[0][0] # 部門
    r2 = a1[0][1] # 科目
    if str.isdigit(r2):
        key2 = int(r2)
    else:
        key2 = 5000000 + ord(r2[0])
    return (key1,key2)
def write_header2(row):
    ce = sh.cell(row=row, column=1)
    ce.value = '部門'
    ce = sh.cell(row=row, column=2)
    ce.value = '科目'
    ce:Cell = sh.cell(row=row, column=3)
    ce.value = '金額'
    ce = sh.cell(row=row, column=4)
    ce.value = '正航的金額'
    

# 準備 acc dict 用
from OneDataOfAcc import OneDataOfAcc
def fn_group3(a1: OneDataOfAcc):    
    return (a1.department, a1.subject) # 部門，科目
def fn_toDict3(a1: t.Tuple[t.Tuple[str,str],lq.Linque]):
    print(a1[0])
    return { a1[0] : a1[1].sum(lambda a2: a2.money)}    
dict_acc = {}
for a1 in lq.linque(data_acc).group(fn_group3):
    dict_acc[a1[0]] = a1[1].sum(lambda a2: a2.money)
    pass
# print(dict_acc)
   
write_header2(row)
row += 1

for a1 in lq.linque(data_keyin_sunday).concat(data_keyin_transfer).group(fn_group2).sort(fn_sort2):
    ce = sh.cell(row=row, column=1)
    ce.value = a1[0][0]
    ce = sh.cell(row=row, column=2)
    ce.value = a1[0][1]
    ce:Cell = sh.cell(row=row, column=3)
    money1 = a1[1].sum(lambda a2: a2.money)
    ce.value = money1
    ce.number_format = '#,##0'
    
    # 比對 acc
    key = a1[0]
    acc = dict_acc.get(key, None)
    if acc is not None:
        ce = sh.cell(row=row, column=4)
        ce.value = acc
        ce.number_format = '#,##0'
        if acc == money1:
            ce.font = openpyxl.styles.Font(color='008800')
        
    row += 1

# 剩下沒比較的 acc
r1 = lq.linque(data_keyin_sunday).concat(data_keyin_transfer).select(fn_group2).distinct().to_list()
def fn_where4(a1):
    return a1 not in r1 and a1[1][0] == '4'
def fn_sort4(a1):
    return a1

# for each dict_acc
for a1 in lq.linque(dict_acc).where(fn_where4).sort(fn_sort4):
    money = dict_acc[a1]    
    
    ce:Cell = sh.cell(row=row, column=1)
    ce.value = a1[0]
    ce:Cell = sh.cell(row=row, column=2)
    ce.value = a1[1]
    ce:Cell = sh.cell(row=row, column=4)
    ce.value = money
    row += 1
    
wb.save('程式測試_比較.xlsx')
wb.close()