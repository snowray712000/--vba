from OnePerson import OnePerson


import linque as lq
import openpyxl


import typing as t


def get_person_and_to_dict(path: str)->t.Dict[t.Union[str,int],OnePerson]:
    ''' 只讀 id2, name, no. 是為了給 240110b 用的 '''
    wbPerson = openpyxl.load_workbook(path)
    sh = wbPerson.active
    rowLast = sh.max_row
    while True and rowLast > 0:
        if sh.cell(row=rowLast, column=1).value is not None:
            break
        rowLast -= 1
    persons: t.List[OnePerson] = lq.linq(range(2, rowLast+1)).select(lambda a1: OnePerson.generate_from_excel(sh, a1)).to_list()
    # print(persons[-1])
    # print(len(persons))

    dictPerson: t.Dict[t.Union[str,int],OnePerson] = {}
    for a1 in persons:
        dictPerson[a1.name] = a1
        if a1.no is not None:
            dictPerson[a1.no] = a1
    # assert dictPerson.get('不具名', None).no == 98
    # assert dictPerson.get(98, None).name == '不具名'
    return dictPerson