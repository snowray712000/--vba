from OneKeyin import OneKeyin


import linque as lq
import openpyxl


import typing as t
from datetime import datetime


def get_all_data_of_keyin(filename: str)-> t.Tuple[t.List[OneKeyin], t.List[OneKeyin]]:
    bk: openpyxl.Workbook = openpyxl.load_workbook('2023 奉獻輸入資料.xlsx')
    sh = bk['輸入原始資料 2023']

    def get_this_sheet(sh)->t.List[OneKeyin]:
        max_row = sh.max_row
        def fn_1(row: int)->bool:
            ''' column B is not None ... 應該是日期 '''
            return sh.cell(row=row, column=2).value is not None
        def fn_select_to_keyin(a1: int)->OneKeyin:
            # 2 日期 3 姓名 4 金額 5 主科目 6 次科目 7 備註
            date = sh.cell(row=a1, column=2).value
            date = datetime.strptime(date, '%Y/%m/%d')
            who = str(sh.cell(row=a1, column=3).value) # 有些格式會不小心變成int, 就影響了後面
            money = sh.cell(row=a1, column=4).value
            subject1 = sh.cell(row=a1, column=5).value
            subject2 = sh.cell(row=a1, column=6).value
            memo = sh.cell(row=a1, column=7).value

            if subject1 is not None: # trim
                # int to string
                subject1 = str(subject1)
                subject1 = subject1.strip()
            if subject2 is not None: # trim
                subject2 = str(subject2)
                subject2 = subject2.strip()

            return OneKeyin(money, date, subject1, subject2, who, memo)
        return lq.linque(range(2, max_row+1)).where(fn_1).select(fn_select_to_keyin).to_list()

    return get_this_sheet(bk['輸入原始資料 2023']), get_this_sheet(bk['輸入原始資料_2023_轉帳'])