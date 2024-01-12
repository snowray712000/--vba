from OneKeyin import OneKeyin
from OnePerson import OnePerson
from openpyxl.worksheet.worksheet import Worksheet
import linque as lq
import typing as t



def export_and_upload_xlsx(shRe: Worksheet,shKeyins: Worksheet, data: t.List[OneKeyin], cntTake: int, tp: t.Literal['主日','轉帳'],dictPerson: t.Dict[t.Union[int,str],OnePerson]) -> None:
    if data is not None:
        # sheet name set Worksheet
        shRe.title = "Worksheet"

        # title 
        # 會友編號/奉獻袋號/身份證 (非會友請留空)	*奉獻者姓名	*奉獻日期 (年/月/日)	*奉獻月份 (選項: 1-12)	*金額	*收取方式 (選項: 金/支票)	*幣別 (請參照系統設定)	*匯率	支票號碼	支票到期日 (年/月/日)	*堂數	*奉獻類別	奉獻子類別	奉獻對象	郵寄郵區	郵寄地址	備註
        shRe.cell(row=1, column=1).value = '會友編號/奉獻袋號/身份證 (非會友請留空)'
        shRe.cell(row=1, column=2).value = '*奉獻者姓名'
        shRe.cell(row=1, column=3).value = '*奉獻日期 (年/月/日)'
        shRe.cell(row=1, column=4).value = '*奉獻月份 (選項: 1-12)'
        shRe.cell(row=1, column=5).value = '*金額'
        shRe.cell(row=1, column=6).value = '*收取方式 (選項: 金/支票)'
        shRe.cell(row=1, column=7).value = '*幣別 (請參照系統設定)'
        shRe.cell(row=1, column=8).value = '*匯率'
        shRe.cell(row=1, column=9).value = '支票號碼'
        shRe.cell(row=1, column=10).value = '支票到期日 (年/月/日)'
        shRe.cell(row=1, column=11).value = '*堂數'
        shRe.cell(row=1, column=12).value = '*奉獻類別'
        shRe.cell(row=1, column=13).value = '奉獻子類別'
        shRe.cell(row=1, column=14).value = '奉獻對象'
        shRe.cell(row=1, column=15).value = '郵寄郵區'
        shRe.cell(row=1, column=16).value = '郵寄地址'
        shRe.cell(row=1, column=17).value = '備註'


        def fn_where1(a1: OneKeyin) -> bool:
            return a1.isUpload == False and a1.subject1str != '其他' and a1.subject1str != '代轉奉獻'
        dataNoUpdateNotOtherNotLoveTransfer: t.List[OneKeyin] = lq.linq(data).where(fn_where1).take(cntTake).to_list()
        # print(len(dataNoUpdateNotOtherNotLoveTransfer))
        # print(dataNoUpdateNotOtherNotLoveTransfer[:3])

        row = 2
        # print(dictPerson)
        for a1 in dataNoUpdateNotOtherNotLoveTransfer:
            user = dictPerson.get(a1.who, None)
            if user is None:
                print(type(a1.who))
                print(f'{a1.who} not found')
                continue

            # id2:1, name:2
            shRe.cell(row=row, column=1).value = user.id2 # 一定是字串
            shRe.cell(row=row, column=1).number_format = '000000000000000'
            shRe.cell(row=row, column=2).value = user.name

            # date:3, month:4, money:5, subject1:12, subject2: 13, memo: 17
            shRe.cell(row=row, column=3).value = a1.date # 2020/12/1 格式
            shRe.cell(row=row, column=3).number_format = 'yyyy/mm/dd'
            shRe.cell(row=row, column=4).value = a1.date.month
            shRe.cell(row=row, column=5).value = a1.money
            shRe.cell(row=row, column=12).value = a1.subject1str
            if a1.subject2str != '':
                shRe.cell(row=row, column=13).value = a1.subject2str
            if a1.memo is not None:
                shRe.cell(row=row, column=17).value = a1.memo

            # 堂數: 11 (目前就只有二種，這字要與奉獻系統一致。)
            tp2: t.Literal['雙福主日','其他'] = '雙福主日' if tp == '主日' else '其它'
            shRe.cell(row=row, column=11).value = tp2

            # Contant: 
            shRe.cell(row=row, column=6).value = '現金'
            shRe.cell(row=row, column=7).value = 'NT'
            shRe.cell(row=row, column=8).value = 1 # 匯率

            a1.isUpload = True

            row += 1

        print('done export sheet')

        for a1 in dataNoUpdateNotOtherNotLoveTransfer:
            a1.setUploadCell(shKeyins)