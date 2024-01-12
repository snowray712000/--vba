from OneDataOfAcc import OneDataOfAcc
from OneKeyin import OneKeyin


import linque as lq
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


import re
import typing as t


def write_transfer_sheet(sh: Worksheet,data_not_sunday_acc: t.List[OneDataOfAcc], deta_not_sunday_keyin: t.List[OneKeyin] ):
    ''' 被 231230.py 呼叫 '''
    def write_header(sh: Worksheet):
        sh.cell(1, 1).value = '傳票號碼'
        sh.cell(1, 2).value = '日期'
        sh.cell(1, 3).value = '科目'
        sh.cell(1, 4).value = '部門'
        sh.cell(1, 5).value = '金額'
        sh.cell(1, 6).value = '摘要'
        sh.cell(1, 8).value = '日期'
        sh.cell(1, 7).value = ''
        sh.cell(1, 9).value = '科目'
        sh.cell(1, 10).value = '部門'
        sh.cell(1, 11).value = '金額'
        sh.cell(1, 12).value = '摘要'
        sh.cell(1, 13).value = '奉獻者' # 要比較奉獻者正確嗎

    # 取得所有 date
    dates = lq.linque(data_not_sunday_acc).select(lambda x: x.date).concat(lq.linque(deta_not_sunday_keyin).select(lambda x: x.date)).distinct().sort().to_list()

    def write_data(sh: Worksheet):
        def _chk_if_acc_in_keyins(acc: OneDataOfAcc, keyins: t.List[OneKeyin])->bool:
            for keyin in keyins:
                if acc.date == keyin.date and acc.money == keyin.money and acc.subject == keyin.subjectNumber[0] and acc.department == keyin.subjectNumber[1]:
                    return True
            return False
        def _chk_if_keyin_in_accs(keyin: OneKeyin, accs: t.List[OneDataOfAcc])->bool:
            for acc in accs:
                if acc.date == keyin.date and acc.money == keyin.money and acc.subject == keyin.subjectNumber[0] and acc.department == keyin.subjectNumber[1]:
                    return True
            return False
        def _chk_keyin_name_in_accs(keyin: OneKeyin, accs: t.List[OneDataOfAcc])->bool:
            if '尾碼' in keyin._who:
                # `尾碼07321` 這種格式時，要取出 07321。
                who = re.search(r'尾碼(\d+)', keyin._who).group(1)

                for acc in accs:
                    if acc.date == keyin.date and acc.money == keyin.money and acc.subject == keyin.subjectNumber[0] and acc.department == keyin.subjectNumber[1] and who in acc.memo:
                        return True
                return False
            else:
                # `阿張三` 或 `16` 這種格式的
                for acc in accs:
                    if acc.date == keyin.date and acc.money == keyin.money and acc.subject == keyin.subjectNumber[0] and acc.department == keyin.subjectNumber[1] and keyin._who in acc.memo:
                        return True
                return False
        row = 2
        for date in dates:
            # 取得這個日期的 acc 與 keyin
            acc2 = lq.linque(data_not_sunday_acc).where(lambda x: x.date == date).to_list()
            keyin2 = lq.linque(deta_not_sunday_keyin).where(lambda x: x.date == date).to_list()

            max_row = max(len(acc2), len(keyin2))
            for i in range(max_row):
                acc3: OneDataOfAcc = acc2[i] if i < len(acc2) else None
                keyin3: OneKeyin = keyin2[i] if i < len(keyin2) else None

                if acc3 is not None:
                    sh.cell(row, 1).value = acc3.subpoena
                    sh.cell(row, 2).value = acc3.date
                    sh.cell(row, 3).value = acc3.subject
                    sh.cell(row, 4).value = acc3.department
                    sh.cell(row, 5).value = acc3.money
                    sh.cell(row, 6).value = acc3.memo

                    if _chk_if_acc_in_keyins(acc3, keyin2):
                        sh.cell(row, 5).font = openpyxl.styles.Font(color='008800')
                if keyin3 is not None:
                    sh.cell(row, 8).value = keyin3.date
                    sh.cell(row, 9).value = keyin3.subjectNumber[0]
                    sh.cell(row, 10).value = keyin3.subjectNumber[1]
                    sh.cell(row, 11).value = keyin3.money
                    sh.cell(row, 12).value = keyin3.memo
                    sh.cell(row, 13).value = keyin3._who

                    if _chk_if_keyin_in_accs(keyin3, acc2):
                        sh.cell(row, 11).font = openpyxl.styles.Font(color='008800')
                    if _chk_keyin_name_in_accs(keyin3, acc2):
                        sh.cell(row, 13).font = openpyxl.styles.Font(color='008800')
                row += 1
            row += 1 # 空一行




    # set date format
    def set_date_column_format(sh: Worksheet):
        for i in range(1, sh.max_row+1):
            cell = sh.cell(i, 2)
            cell.number_format = 'm/d'
            cell = sh.cell(i, 8)
            cell.number_format = 'm/d'
    def set_money_column_format(sh: Worksheet):
        for i in range(1, sh.max_row+1):
            cell = sh.cell(i, 5)
            cell.number_format = '#,##0'
            cell = sh.cell(i, 11)
            cell.number_format = '#,##0'

    write_header(sh)
    write_data(sh)
    set_date_column_format(sh)
    set_money_column_format(sh)