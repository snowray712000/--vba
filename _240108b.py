'''
- 需新增的會員
    - 上傳的時候，必須要有會員資料，不然無法透過 excel 方式上傳。
    - 載入目前會員。接著讀取所有奉獻keyin資料，最後列出哪些姓名是缺少的。
'''

#%%
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import linque as lq
import typing as t

#%%
def get_keyin_users():    
    wb2 = openpyxl.load_workbook('2023 奉獻輸入資料.xlsx')
    sh2a = wb2['輸入原始資料 2023']
    sh2b = wb2['輸入原始資料_2023_轉帳']
    
    def fn_sort1(a1):    
        if isinstance(a1, int):
            return a1
        if False == str.isdigit(a1):
            return 9999 + 20000*len(a1) + ord(a1[0])
        else:
            return int(a1)
    def fn_select1(a1):
        if isinstance(a1, int):
            return a1
        else:
            if False == str.isdigit(a1):
                return a1
            return int(a1)
    def get_user(sh: Worksheet):
        def fn_where2(r):
            ce = sh.cell(r, 4).value
            if ce is None:
                return False
            if ce == 'x':
                return False
            return True
        return lq.linque(range(2, sh.max_row + 1)).where(fn_where2).select(lambda i: sh.cell(i, 3).value).where(lambda a1: a1 is not None).select(fn_select1).distinct().sort(fn_sort1).to_list()

    r1 = get_user(sh2b)
    r1.extend(get_user(sh2a))
    return lq.linque(r1).distinct().sort(fn_sort1).to_list()

user_keyin = get_keyin_users()
#%%
def get_church_users()->t.List[t.Tuple[str, t.Optional[int]]]:
    wb = openpyxl.load_workbook('雙福教會會友列表2023.xlsx')
    sh = wb.active
    
    def fn_select1(r):
        r1 = sh.cell(r,2).value
        r2 = sh.cell(r,18).value
        if r2 is None:
            return (r1, None)
        else:
            return (r1, int(r2))
    def fn_sort2(a1: t.Tuple[str, t.Optional[int]]):    
        if a1[1] is not None:
            return a1[1]
        else:
            r1 = a1[0]
            return 9999 + 20000 * len(r1) + ord(r1[0])   
    return lq.linque(range(2, sh.max_row + 1)).select(fn_select1).where(lambda a1: a1[0] is not None).sort(fn_sort2).to_list()

user_db = get_church_users()

#%%

wb3 = Workbook()
sh3 = wb3.active

# 建立 user_db dict, 以便知道哪些有用過
user_db_used = {a1[0]: False for a1 in user_db}
print(user_db_used)

# 檢查 user_keyin
user_keyin_used = {a1: False for a1 in user_keyin}
print(user_keyin_used)

# 為了提昇 search 效率，先建 dict
user_db_dict = {a1[0]: a1 for a1 in user_db}
for a1 in user_db:
    if a1[1] is not None:
        user_db_dict[a1[1]] = a1

# 輸出有使用到的
row = 2
for a1 in user_keyin:
    user_keyin_used[a1] = True        
    
    r2 = user_db_dict.get(a1, None)
    if r2 is None:
        sh3.cell(row, 1).value = a1
        ce = sh3.cell(row, 1)
        ce.font = openpyxl.styles.Font(color='FF0000')
    else:
        sh3.cell(row, 1).value = a1
        sh3.cell(row, 2).value = r2[0]
        if r2[1] is not None:
            sh3.cell(row, 3).value = r2[1]
        user_db_used[r2[0]] = True
    row += 1

# 輸出 db 沒有用到的
row += 1
for a1 in user_db_used:
    if user_db_used[a1] == False:
        r1 = user_db_dict.get(a1, None)
        sh3.cell(row, 2).value = r1[0]
        if r1[1] is not None:
            sh3.cell(row, 3).value = r1[1]
        row += 1

wb3.save('程式測試_需新增人員至系統.xlsx')