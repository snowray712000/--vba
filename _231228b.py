#%%

import openpyxl

from get_all_data_of_acc import get_all_data_of_acc
from generate_sheet_of_subpoena import generate_sheet_of_subpoena
from filter_subpoena_sunday_and_income import filter_subpoena_sunday_and_income, filter_subpoena_not_sunday_and_income
from sort_acc import sort_acc

#%
# 取得科目代碼 1-6 開頭的
rows = get_all_data_of_acc('2023傳票_20231216匯出.xlsx')
wb = openpyxl.Workbook()
wb.remove(wb.worksheets[0])
sh = wb.create_sheet('傳票')
generate_sheet_of_subpoena(sh, rows)

#%
# 目的: 為了預備對照奉獻的輸入資料
rows2 = filter_subpoena_sunday_and_income(rows)
generate_sheet_of_subpoena(wb.create_sheet('傳票_主日_收入'), sort_acc(rows2))

#%
rows3 = filter_subpoena_not_sunday_and_income(rows)
generate_sheet_of_subpoena(wb.create_sheet('傳票_非主日_收入'), rows3)

#%

wb.save("程式測試_正航.xlsx")
wb.close()




