#%%

import openpyxl
from generate_sheet_of_keyin import generate_sheet_of_keyin
from get_all_data_of_keyin import get_all_data_of_keyin
from grouping_keyin import grouping_keyin
from sort_keyin import sort_keyin
rows, rows2 = get_all_data_of_keyin('2023 奉獻輸入資料.xlsx')

wb = openpyxl.Workbook()
wb.remove(wb.worksheets[0])

sh = wb.create_sheet('keyin_主日')
generate_sheet_of_keyin(sh, rows)

sh = wb.create_sheet('keyin_轉帳')
generate_sheet_of_keyin(sh, rows2)
#%
rows3 = grouping_keyin(rows)
rows3b = sort_keyin(rows3)    
generate_sheet_of_keyin(wb.create_sheet('keyin_主日_統計'), rows3b)

wb.save("程式測試_keyin.xlsx")
wb.close()