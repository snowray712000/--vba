'''
- 產生成為袋號總表。就是幹事用的那個表格。
'''
#%%
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

wb = openpyxl.load_workbook('雙福教會會友列表2023.xlsx')
sh: Worksheet = wb.active

#%%
datas = []
for r in range(2, sh.max_row):
    no = sh.cell(row=r, column=18).value    
    name = sh.cell(row=r, column=2).value
    datas.append( (no, name) )

import linque as lq

r1 = lq.linque(datas).where(lambda x: x[0] is not None).sort(lambda x: int(x[0])).group(lambda x: x[0])

# for a1 in r1:
#     r2: lq.Linque = a1[1]
#     if r2.count() > 1:
#         print(a1[0], r2.to_list())
#     else:
#         print(r2.first())
#%
def get_row_col_of_no(no:int):
    # 1 row = 3 col = 1
    # 20 row = 22 col = 1
    # 21 row = 3 col = 3
    # 40 row = 22 col = 3
    # 41 row = 3 col = 5
    # 60 row = 22 col = 5
    r1 = (no - 1) // 20
    r2 = (no - 1) % 20
    col = r1 * 2 + 1
    row = 3 + r2
    return row, col

wb2 = Workbook()
sh2: Worksheet = wb2.active

r1 = r1.to_list()
for a1 in r1:
    r2: lq.Linque = a1[1]
    no = int ( a1[0] )
    r, c = get_row_col_of_no(no)
    # print(no, r, c)
    
    names = ';'.join(r2.select(lambda x: x[1]).to_list())
    sh2.cell(row=r, column=c).value = int(a1[0])
    
    cell = sh2.cell(row=r, column=c+1)
    cell.value = names
    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    if r2.count() > 1:
        cell.font = openpyxl.styles.Font(color='880000')

# titles, row=2
lastNo = int(r1[-1][0])
rLast, cLast = get_row_col_of_no(lastNo)
for c in range(1, cLast+2):
    cell = sh2.cell(row=2, column=c)
    if c % 2 == 1:
        cell.value = '袋號'
    else:
        cell.value = '姓名'
    
    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
    cell.fill = openpyxl.styles.PatternFill('solid', fgColor='ccffff')
    
# merge cell (column 1 to max_column) of row 1
sh2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cLast+1)
sh2.cell(row=1, column=1).value = '2022 雙福教會奉獻袋號編號表'
sh2.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(horizontal='center')

# grid line
for r in range(1, sh2.max_row+1):
    for c in range(1, cLast+2):
        cell = sh2.cell(row=r, column=c)
        cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))


wb2.save('程式測試_會友列表.xlsx')