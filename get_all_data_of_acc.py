#%%

# get all data
from OneDataOfAcc import OneDataOfAcc


import openpyxl


import typing as t


def get_all_data_of_acc(path: str)->t.List[OneDataOfAcc]:
    """ 分析匯出的傳票，並只保留有效的Row，即科目代碼第1碼只可能是 1-6
    """
    book = openpyxl.load_workbook(path)
    sh = book.worksheets[0]
    # 取得總 row 數
    max_row = sh.max_row
    def valid_row(sh: openpyxl.worksheet.worksheet.Worksheet, row: int)->bool:
        """ 檢查是否為有效的 row 
        - 科目代碼， 第1碼只可能是 1-6
        """
        ce5 = sh.cell(row = row, column=5)
        if ce5.value is None or ce5.value == '':
            return False

        # 科目代碼，第1碼只可能是 1-6
        if ce5.value[0] not in '123456':
            return False
        return True

    rows: t.List[OneDataOfAcc] = []
    # 19 摘要 3 部門 4 傳票號標 5 科目 6 科目名 14 金額
    for i in range(1, max_row+1):
        if valid_row(sh, i) is False: continue

        # JC20230101002 傳票編號 例子， 20230101 可以轉成日期
        one = OneDataOfAcc()

        subpoena = sh.cell(row = i, column=4) # 傳票編號    
        department = sh.cell(row = i, column=3) # 部門編號
        memo = sh.cell(row = i, column=19) # 摘要
        money = sh.cell(row = i, column=14) # 金額
        subject = sh.cell(row = i, column=5) # 科目代碼

        one.set(subject.value, subpoena.value, department.value, memo.value, money.value)
        rows.append(one)
    
    return rows